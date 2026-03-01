import asyncio
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from app.db.session import SessionLocal
from app.models.entities import HSNMaster, Pricing, Product


def _resolve_dataset_dir() -> Path:
    # Allow running the script from any cwd by searching upward for repo dataset folder.
    for parent in Path(__file__).resolve().parents:
        candidate = parent / "dataset"
        if (
            (candidate / "Inventory_Cleaned.xlsx").exists()
            and (candidate / "Stock.xlsx").exists()
            and (candidate / "HSNSAC_MASTER.xlsx").exists()
        ):
            return candidate
    raise FileNotFoundError(
        "Could not locate dataset folder containing Inventory_Cleaned.xlsx, Stock.xlsx, and HSNSAC_MASTER.xlsx"
    )


DATASET_DIR = _resolve_dataset_dir()
INVENTORY_FILE = DATASET_DIR / "Inventory_Cleaned.xlsx"
STOCK_FILE = DATASET_DIR / "Stock.xlsx"
HSN_FILE = DATASET_DIR / "HSNSAC_MASTER.xlsx"


def _clean_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _to_decimal(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        cleaned = str(value).strip().replace(",", "")
        if cleaned == "":
            return None
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None


def _first_decimal(*values) -> Decimal | None:
    for value in values:
        if value is not None:
            return value
    return None


def _parse_tax_percent(value) -> Decimal | None:
    text = _clean_text(value)
    if not text:
        return None
    numbers = [Decimal(match) for match in re.findall(r"\d+(?:\.\d+)?", text)]
    if not numbers:
        return None
    if len(numbers) >= 3:
        summed = numbers[0] + numbers[1]
        if numbers[2] == summed:
            return numbers[2]
    if len(numbers) >= 2:
        return numbers[0] + numbers[1]
    return numbers[0]


def _pct_diff(price: Decimal | None, mrp: Decimal | None) -> Decimal | None:
    if price is None or mrp is None or mrp == 0:
        return None
    return ((price - mrp) / mrp) * Decimal("100")


def _load_hsn_rows() -> dict[str, dict]:
    wb = load_workbook(HSN_FILE, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [_clean_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]

    rows: dict[str, dict] = {}
    for r in range(2, ws.max_row + 1):
        raw = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1) if headers[c - 1]}
        hsn_code = _clean_text(raw.get("HSN Code"))
        if not hsn_code:
            continue
        rows[hsn_code] = {
            "hsn_code": hsn_code,
            "description": _clean_text(raw.get("HSN Name")),
            "gst_percent": _parse_tax_percent(raw.get("Tax")) or Decimal("0"),
        }
    return rows


def _load_inventory_rows() -> dict[str, dict]:
    wb = load_workbook(INVENTORY_FILE, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [_clean_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]

    rows: dict[str, dict] = {}
    for r in range(2, ws.max_row + 1):
        raw = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1) if headers[c - 1]}
        sku = _clean_text(raw.get("SKU"))
        if not sku:
            continue

        rows[sku] = {
            "sku": sku,
            "name": sku,
            "brand": _clean_text(raw.get("Brand")),
            "taxable_value": _to_decimal(raw.get("Taxable Value")),
            "mrp": _to_decimal(raw.get("MRP")),
            "rate_a": _to_decimal(raw.get("RATE-A")),
            "rate_b": _to_decimal(raw.get("Rate-B")),
            "tax_percent": _to_decimal(raw.get("TAX")) or _to_decimal(raw.get("GST")) or Decimal("0"),
        }
    return rows


def _load_stock_lookup() -> dict[str, dict]:
    wb = load_workbook(STOCK_FILE, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [_clean_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]

    lookup: dict[str, dict] = {}
    for r in range(2, ws.max_row + 1):
        raw = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1) if headers[c - 1]}
        sku = _clean_text(raw.get("Item Name(Internal - SKU)"))
        if not sku:
            continue
        lookup[sku] = {
            "hsn_code": _clean_text(raw.get("HSN Number")),
            "display_name": _clean_text(raw.get("Item Display Name")),
            "category": _clean_text(raw.get("Category")),
            "sub_category": _clean_text(raw.get("Sub Category")),
            "brand": _clean_text(raw.get("Brand")),
            "mrp": _to_decimal(raw.get("MRP")),
            "taxable_value": _to_decimal(raw.get("Taxable Value")),
            "base_price": _to_decimal(raw.get("A-Rate Outlets  (Mention Percentage Margin from Base Rate)")),
            "b_class_price": _to_decimal(raw.get("B-Rate Outlets (Mention Percentage Margin from Base Rate)")),
            "c_class_price": _to_decimal(raw.get("B2C Rate  (Mention Percentage Margin from Base Rate)")),
            "tax_percent": _to_decimal(raw.get("CGST(Auto)")) * Decimal("2")
            if _to_decimal(raw.get("CGST(Auto)")) is not None
            else None,
        }
    return lookup


async def import_products() -> None:
    hsn_rows = _load_hsn_rows()
    inventory_rows = _load_inventory_rows()
    stock_lookup = _load_stock_lookup()

    hsn_created = 0
    hsn_updated = 0
    created = 0
    updated = 0
    pricing_created = 0
    pricing_updated = 0

    async with SessionLocal() as session:
        existing_hsn = {
            row.hsn_code: row
            for row in (
                await session.execute(select(HSNMaster).where(HSNMaster.hsn_code.in_(list(hsn_rows.keys()))))
            ).scalars()
        }
        for hsn_code, row in hsn_rows.items():
            current = existing_hsn.get(hsn_code)
            if current is None:
                current = HSNMaster(
                    hsn_code=hsn_code,
                    description=row["description"],
                    gst_percent=row["gst_percent"],
                    is_active=True,
                )
                session.add(current)
                existing_hsn[hsn_code] = current
                hsn_created += 1
            else:
                current.description = row["description"] or current.description
                current.gst_percent = row["gst_percent"]
                current.is_active = True
                hsn_updated += 1

        await session.flush()
        hsn_id_by_code = {code: row.id for code, row in existing_hsn.items()}

        existing_products = {
            row.sku: row
            for row in (await session.execute(select(Product).where(Product.sku.in_(list(inventory_rows.keys()))))).scalars()
        }

        for sku, row in inventory_rows.items():
            stock_row = stock_lookup.get(sku, {})
            base_price = _first_decimal(stock_row.get("base_price"), row.get("rate_a"), row.get("taxable_value"), Decimal("0"))
            tax_percent = _first_decimal(stock_row.get("tax_percent"), row.get("tax_percent"), Decimal("0"))
            values = {
                "sku": sku,
                "name": row["name"],
                "display_name": stock_row.get("display_name"),
                "brand": stock_row.get("brand") or row["brand"],
                "category": stock_row.get("category"),
                "sub_category": stock_row.get("sub_category"),
                "hsn_id": hsn_id_by_code.get(stock_row.get("hsn_code")) if stock_row.get("hsn_code") else None,
                "unit": "PCS",
                "base_price": base_price,
                "tax_percent": tax_percent,
                "is_active": True,
            }

            existing = existing_products.get(sku)
            if existing is None:
                session.add(Product(**values))
                created += 1
            else:
                for key, value in values.items():
                    if key == "sku":
                        continue
                    setattr(existing, key, value)
                updated += 1

        await session.flush()
        products_by_sku = {
            row.sku: row
            for row in (await session.execute(select(Product).where(Product.sku.in_(list(inventory_rows.keys()))))).scalars()
        }

        product_ids = [product.id for product in products_by_sku.values()]
        existing_pricing = {
            row.product_id: row
            for row in (
                await session.execute(select(Pricing).where(Pricing.product_id.in_(product_ids)))
            ).scalars()
        }

        for sku, product in products_by_sku.items():
            inventory_row = inventory_rows[sku]
            stock_row = stock_lookup.get(sku, {})
            mrp = _first_decimal(inventory_row.get("mrp"), stock_row.get("mrp"), Decimal("0")) or Decimal("0")
            cost_price = _first_decimal(inventory_row.get("taxable_value"), stock_row.get("taxable_value"), product.base_price) or Decimal(
                "0"
            )
            a_price = _first_decimal(stock_row.get("base_price"), inventory_row.get("rate_a"), product.base_price) or Decimal("0")
            b_price = _first_decimal(stock_row.get("b_class_price"), inventory_row.get("rate_b"), a_price) or Decimal("0")
            c_price = _first_decimal(stock_row.get("c_class_price"), mrp, b_price) or Decimal("0")

            values = {
                "product_id": product.id,
                "mrp": mrp,
                "cost_price": cost_price,
                "a_class_price": a_price,
                "b_class_price": b_price,
                "c_class_price": c_price,
                "pct_diff_a_mrp": _pct_diff(a_price, mrp),
                "pct_diff_b_mrp": _pct_diff(b_price, mrp),
                "pct_diff_c_mrp": _pct_diff(c_price, mrp),
                "is_active": True,
            }

            current = existing_pricing.get(product.id)
            if current is None:
                session.add(Pricing(**values))
                pricing_created += 1
            else:
                for key, value in values.items():
                    if key == "product_id":
                        continue
                    setattr(current, key, value)
                pricing_updated += 1

        await session.commit()

    print(
        "Import complete. "
        f"hsn_created={hsn_created}, hsn_updated={hsn_updated}, "
        f"products_created={created}, products_updated={updated}, "
        f"pricing_created={pricing_created}, pricing_updated={pricing_updated}, "
        f"total_skus={len(inventory_rows)}"
    )


if __name__ == "__main__":
    asyncio.run(import_products())
